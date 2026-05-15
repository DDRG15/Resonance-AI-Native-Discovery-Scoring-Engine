"""
gema_industrial.py — GEMA Industrial Multi-Source Scraper
Scrapes up to 100 jobs from 8 job boards, extracts via NLP, writes multi-tab Excel.
"""

import argparse
import asyncio
import gzip
import json
import re
import random
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import date
from html.parser import HTMLParser
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, ".")
from dotenv import load_dotenv
load_dotenv()

from nlp_engine import (
    extract_jobs_from_text,
    reset_rate_limit_flags,
    generate_ephemeral_profile,
    extract_cv_text,
    _build_extraction_prompt,
)

# =============================================================================
# Constants
# =============================================================================

XLSX_PATH  = Path(__file__).parent / "GEMA_JOB_TRACKER.xlsx"
BATCH_SIZE = 5
MAX_JOBS   = 100
TODAY      = date.today().isoformat()

WELLFOUND_WALK_MAX_BREADTH = 200  # max items traversed per level in _walk()
CIRCUIT_BREAKER_MAX_ERRORS = 3    # consecutive NLP batch failures before aborting a source

COLUMNS = [
    "DATE_FETCHED", "MATCH_SCORE", "JOB_TITLE", "COMPANY",
    "SALARY", "LOCATION_STRICTNESS", "TECH_STACK", "APPLY_LINK", "NOTES",
]

BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",
    "Connection": "keep-alive",
    "DNT": "1",
    "Upgrade-Insecure-Requests": "1",
}

JSON_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate",
}

# =============================================================================
# Utilities
# =============================================================================

class _Strip(HTMLParser):
    def __init__(self):
        super().__init__()
        self._p: list[str] = []

    def handle_data(self, d: str):
        t = d.strip()
        if t:
            self._p.append(t)

    def get_text(self) -> str:
        return "\n".join(self._p)


def html_to_text(h: str) -> str:
    p = _Strip()
    p.feed(h.encode("utf-8", errors="replace").decode("utf-8", errors="replace"))
    return p.get_text()


def safe_get(url: str, headers: dict = None, timeout: int = 20) -> str:
    req = urllib.request.Request(url, headers=headers or BROWSER_HEADERS)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            data = r.read()
            enc = r.headers.get("Content-Encoding", "")
            if "gzip" in enc:
                data = gzip.decompress(data)
            return data.decode("utf-8", errors="replace")
    except Exception as e:
        print(f"  [WARN] GET failed {url}: {type(e).__name__}")
        return ""


def safe_get_json(url: str, headers: dict = None, timeout: int = 20) -> dict | list:
    raw = safe_get(url, headers=headers or JSON_HEADERS, timeout=timeout)
    if not raw:
        return {}
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return {}


def build_text_block(i: int, title: str = "", company: str = "", url: str = "",
                     description: str = "", salary: str = "", location: str = "") -> str:
    parts = [f"Job {i}:"]
    if title:
        parts.append(f"Title: {title}")
    if company:
        parts.append(f"Company: {company}")
    if location:
        parts.append(f"Location: {location}")
    if salary:
        parts.append(f"Salary: {salary}")
    if url:
        parts.append(f"URL: {url}")
    if description:
        parts.append(f"\n{description[:3000]}")
    return "\n".join(parts)


# =============================================================================
# Batch NLP extraction
# =============================================================================

def _fmt_salary(job) -> str:
    if job.salary_min is None and job.salary_max is None:
        return "Not disclosed"
    curr = job.currency or "USD"
    lo = f"{job.salary_min:,}" if job.salary_min else "?"
    hi = f"{job.salary_max:,}" if job.salary_max else "?"
    if job.salary_min and job.salary_max and job.salary_min != job.salary_max:
        return f"{curr} {lo} - {hi}"
    return f"{curr} {lo}"


def _fmt_tech(job) -> str:
    return ", ".join(job.required_tech) if job.required_tech else "Not specified"


def extract_and_build_rows(
    raw_jobs: list[dict],
    source_url: str,
    seen_urls: Optional[set] = None,
    extraction_prompt: Optional[str] = None,
) -> list[dict]:
    """
    NLP-extracts structured rows from raw scraped job dicts.

    seen_urls: shared set across sources — mutated in-place for cross-source dedup.
    extraction_prompt: personalised prompt from generate_ephemeral_profile(); falls
        back to the generic _EXTRACTION_SYSTEM_PROMPT inside extract_jobs_from_text().
    Circuit breaker: aborts this source after CIRCUIT_BREAKER_MAX_ERRORS consecutive
        NLP batch failures to prevent burning API quota on a broken feed.
    """
    if seen_urls is None:
        seen_urls = set()
    rows = []
    consecutive_errors = 0

    for batch_start in range(0, len(raw_jobs), BATCH_SIZE):
        if consecutive_errors >= CIRCUIT_BREAKER_MAX_ERRORS:
            print(
                f"  [CIRCUIT BREAKER] {CIRCUIT_BREAKER_MAX_ERRORS} consecutive NLP failures "
                f"— aborting this source to preserve API quota."
            )
            break

        batch = [
            rj for rj in raw_jobs[batch_start: batch_start + BATCH_SIZE]
            if rj.get("url", "") not in seen_urls
        ]
        if not batch:
            continue

        blocks = []
        for local_i, rj in enumerate(batch, batch_start + 1):
            block = build_text_block(
                local_i,
                title=rj.get("title", ""),
                company=rj.get("company", ""),
                url=rj.get("url", source_url),
                description=rj.get("description", ""),
                salary=rj.get("salary", ""),
                location=rj.get("location", ""),
            )
            blocks.append(block)

        combined = "\n\n" + ("\n\n" + "-" * 40 + "\n\n").join(blocks)
        try:
            result = extract_jobs_from_text(
                combined,
                source_url=source_url,
                log_callback=None,
                extraction_prompt=extraction_prompt,
            )
            consecutive_errors = 0
        except Exception as e:
            print(f"  [ERROR] NLP batch failed: {e}")
            consecutive_errors += 1
            continue

        for job in result.jobs:
            url = job.source_url or source_url
            if url in seen_urls:
                continue
            seen_urls.add(url)
            rows.append({
                "DATE_FETCHED":        TODAY,
                "MATCH_SCORE":         job.cv_match_score,
                "JOB_TITLE":           job.title,
                "COMPANY":             job.company,
                "SALARY":              _fmt_salary(job),
                "LOCATION_STRICTNESS": job.location_strictness or "Unknown",
                "TECH_STACK":          _fmt_tech(job),
                "APPLY_LINK":          url,
                "NOTES":               job.location_notes or "",
            })
            print(f"  Extracted: {job.title} at {job.company} [{job.cv_match_score:.0%}]")

    return rows


# =============================================================================
# SCRAPER 1: HackerNews Jobs
# =============================================================================

def scrape_hackernews(max_jobs: int = MAX_JOBS) -> list[dict]:
    print("\n[GEMA] Scraping HackerNews Jobs...")

    class _HNParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.jobs: list[dict] = []
            self._in_titleline = False
            self._cur_url = ""
            self._cur_texts: list[str] = []

        def handle_starttag(self, tag, attrs):
            ad = dict(attrs)
            if tag == "span" and ad.get("class") == "titleline":
                self._in_titleline = True
                self._cur_url = ""
                self._cur_texts = []
            elif self._in_titleline and tag == "a" and not self._cur_url:
                self._cur_url = ad.get("href", "")

        def handle_data(self, data):
            if self._in_titleline and data.strip():
                self._cur_texts.append(data.strip())

        def handle_endtag(self, tag):
            if tag == "span" and self._in_titleline:
                text = " ".join(t for t in self._cur_texts if t)
                if text and self._cur_url:
                    self.jobs.append({"title": text, "url": self._cur_url})
                self._in_titleline = False

    raw_jobs: list[dict] = []
    seen: set[str] = set()
    url = "https://news.ycombinator.com/jobs"

    for _ in range(4):
        if len(raw_jobs) >= max_jobs:
            break
        html = safe_get(url)
        if not html:
            break

        parser = _HNParser()
        parser.feed(html)

        for job in parser.jobs:
            job_url = job["url"]
            if not job_url.startswith("http"):
                job_url = "https://news.ycombinator.com/" + job_url.lstrip("/")
            if job_url not in seen and len(raw_jobs) < max_jobs:
                seen.add(job_url)
                raw_jobs.append({
                    "title": job["title"],
                    "url": job_url,
                    "company": "",
                    "description": "",
                })

        next_match = re.search(r'href="(/jobs\?next=\d+)"', html)
        if next_match:
            url = "https://news.ycombinator.com" + next_match.group(1)
            time.sleep(random.uniform(1.5, 2.5))
        else:
            break

    print(f"  Found {len(raw_jobs)} HN jobs")
    return raw_jobs[:max_jobs]


# =============================================================================
# SCRAPER 2: WorkingNomads (public API)
# =============================================================================

def scrape_workingnomads(max_jobs: int = MAX_JOBS) -> list[dict]:
    print("\n[GEMA] Scraping WorkingNomads API...")
    raw_jobs: list[dict] = []
    seen: set[str] = set()

    for cat in ["development", "devops-sysadmin", "security"]:
        if len(raw_jobs) >= max_jobs:
            break
        data = safe_get_json(
            f"https://www.workingnomads.com/api/exposed_jobs/?category={cat}"
        )
        jobs = data if isinstance(data, list) else data.get("results", [])

        for j in jobs:
            if len(raw_jobs) >= max_jobs:
                break
            job_url = j.get("url") or j.get("apply_url") or ""
            if not job_url or job_url in seen:
                continue
            seen.add(job_url)
            desc = j.get("description") or ""
            raw_jobs.append({
                "title":       j.get("title", ""),
                "company":     j.get("company_name") or j.get("company") or "",
                "url":         job_url,
                "location":    j.get("location") or j.get("region") or "",
                "description": html_to_text(desc) if "<" in desc else desc,
            })

        time.sleep(0.5)

    print(f"  Found {len(raw_jobs)} WorkingNomads jobs")
    return raw_jobs[:max_jobs]


# =============================================================================
# SCRAPER 3: Himalayas (public API)
# =============================================================================

def scrape_himalayas(max_jobs: int = MAX_JOBS) -> list[dict]:
    print("\n[GEMA] Scraping Himalayas API...")
    raw_jobs: list[dict] = []
    seen: set[str] = set()
    page = 1

    while len(raw_jobs) < max_jobs and page <= 5:
        data = safe_get_json(
            f"https://himalayas.app/jobs/api?page={page}&quantity=50"
        )
        jobs = data.get("jobs", [])
        if not jobs:
            break

        for j in jobs:
            if len(raw_jobs) >= max_jobs:
                break

            # Resolve job URL
            job_url = j.get("applicationLink") or j.get("applyUrl") or ""
            slug = j.get("slug", "")
            if not job_url and slug:
                job_url = f"https://himalayas.app/jobs/{slug}"
            if not job_url or job_url in seen:
                continue
            seen.add(job_url)

            company_obj = j.get("company", {})
            company = (company_obj.get("name", "") if isinstance(company_obj, dict)
                       else str(company_obj))

            sal_obj = j.get("salary", {})
            salary_str = ""
            if isinstance(sal_obj, dict) and sal_obj.get("min"):
                lo, hi = sal_obj["min"], sal_obj.get("max")
                salary_str = (f"USD {lo:,} - {hi:,}" if hi and hi != lo
                              else f"USD {lo:,}+")

            desc = j.get("description") or j.get("summary") or ""
            raw_jobs.append({
                "title":       j.get("title", ""),
                "company":     company,
                "url":         job_url,
                "location":    j.get("location") or j.get("locationPolicy") or "",
                "salary":      salary_str,
                "description": html_to_text(desc) if "<" in desc else desc,
            })

        page += 1
        time.sleep(0.5)

    print(f"  Found {len(raw_jobs)} Himalayas jobs")
    return raw_jobs[:max_jobs]


# =============================================================================
# SCRAPER 4: Remotivated (HTML)
# =============================================================================

def scrape_remotivated(max_jobs: int = MAX_JOBS) -> list[dict]:
    print("\n[GEMA] Scraping Remotivated...")
    raw_jobs: list[dict] = []
    seen: set[str] = set()

    html = safe_get("https://remotivated.com/jobs")
    if not html:
        print("  [WARN] Remotivated: no response")
        return []

    # Try JSON-LD structured data first
    for jld_raw in re.findall(
        r'<script[^>]+type="application/ld\+json"[^>]*>(.*?)</script>', html, re.DOTALL
    ):
        try:
            items = json.loads(jld_raw)
            if not isinstance(items, list):
                items = [items]
            for item in items:
                if item.get("@type") != "JobPosting" or len(raw_jobs) >= max_jobs:
                    continue
                job_url = item.get("url", "")
                if not job_url or job_url in seen:
                    continue
                seen.add(job_url)
                desc = item.get("description", "")
                org = item.get("hiringOrganization") or {}
                raw_jobs.append({
                    "title":       item.get("title", ""),
                    "company":     org.get("name", "") if isinstance(org, dict) else "",
                    "url":         job_url,
                    "location":    "",
                    "description": html_to_text(desc) if "<" in desc else desc,
                })
        except Exception:
            pass

    # Fallback: find /jobs/* links in HTML
    if len(raw_jobs) < max_jobs:
        for link in re.findall(r'href="(/jobs/[^"?#]+)"', html):
            if len(raw_jobs) >= max_jobs:
                break
            full_url = f"https://remotivated.com{link}"
            if full_url in seen:
                continue
            seen.add(full_url)
            slug = link.rstrip("/").split("/")[-1]
            raw_jobs.append({
                "title":       slug.replace("-", " ").title(),
                "company":     "",
                "url":         full_url,
                "description": "",
            })

    print(f"  Found {len(raw_jobs)} Remotivated jobs")
    return raw_jobs[:max_jobs]


# =============================================================================
# SCRAPER 5: PostHog Cool Tech Jobs (HTML)
# =============================================================================

def scrape_posthog(max_jobs: int = MAX_JOBS) -> list[dict]:
    print("\n[GEMA] Scraping PostHog Cool Tech Jobs...")
    raw_jobs: list[dict] = []
    seen: set[str] = set()

    html = safe_get("https://posthog.com/cool-tech-jobs")
    if not html:
        print("  [WARN] PostHog: no response")
        return []

    # Find external job links (not posthog.com itself)
    JOB_RE = re.compile(
        r"(engineer|developer|backend|frontend|devops|sre|platform|data|security|infra)", re.I
    )
    for link_url, link_text in re.findall(r'href="(https?://[^"]+)"[^>]*>([^<]{3,120})<', html):
        if len(raw_jobs) >= max_jobs:
            break
        if "posthog.com" in link_url:
            continue
        if not JOB_RE.search(link_text + " " + link_url):
            continue
        if link_url in seen:
            continue
        seen.add(link_url)
        raw_jobs.append({
            "title":       link_text.strip(),
            "company":     "",
            "url":         link_url,
            "description": f"Listed on PostHog cool-tech-jobs. Role: {link_text.strip()}",
        })

    print(f"  Found {len(raw_jobs)} PostHog jobs")
    return raw_jobs[:max_jobs]


# =============================================================================
# SCRAPER 6: Greenhouse Careers (HTML + API fallback)
# =============================================================================

def scrape_greenhouse(max_jobs: int = MAX_JOBS) -> list[dict]:
    print("\n[GEMA] Scraping Greenhouse Careers...")
    raw_jobs: list[dict] = []
    seen: set[str] = set()

    # Try their public board API for Greenhouse's own company jobs
    data = safe_get_json(
        "https://boards-api.greenhouse.io/v1/boards/greenhouse/jobs?content=true"
    )
    for j in data.get("jobs", [])[:max_jobs]:
        job_url = j.get("absolute_url", "")
        if not job_url or job_url in seen:
            continue
        seen.add(job_url)
        desc = j.get("content", "")
        loc = j.get("location", {})
        raw_jobs.append({
            "title":       j.get("title", ""),
            "company":     "Greenhouse",
            "url":         job_url,
            "location":    loc.get("name", "") if isinstance(loc, dict) else "",
            "description": html_to_text(desc) if "<" in desc else desc,
        })

    # Also scrape the careers page for additional listings
    if len(raw_jobs) < max_jobs:
        html = safe_get("https://www.greenhouse.com/careers/opportunities")
        if html:
            for jld_raw in re.findall(
                r'<script[^>]+type="application/ld\+json"[^>]*>(.*?)</script>',
                html, re.DOTALL
            ):
                try:
                    items = json.loads(jld_raw)
                    if not isinstance(items, list):
                        items = [items]
                    for item in items:
                        if item.get("@type") != "JobPosting":
                            continue
                        job_url = item.get("url", "")
                        if not job_url or job_url in seen or len(raw_jobs) >= max_jobs:
                            continue
                        seen.add(job_url)
                        desc = item.get("description", "")
                        org = item.get("hiringOrganization") or {}
                        raw_jobs.append({
                            "title":       item.get("title", ""),
                            "company":     org.get("name", "Greenhouse") if isinstance(org, dict) else "Greenhouse",
                            "url":         job_url,
                            "location":    "",
                            "description": html_to_text(desc) if "<" in desc else desc,
                        })
                except Exception:
                    pass

    print(f"  Found {len(raw_jobs)} Greenhouse jobs")
    return raw_jobs[:max_jobs]


# =============================================================================
# SCRAPER 7: Wellfound (Playwright — stealth mode)
# =============================================================================

async def _playwright_wellfound(max_jobs: int) -> list[dict]:
    from playwright.async_api import async_playwright  # type: ignore

    raw_jobs: list[dict] = []
    seen: set[str] = set()

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-infobars",
                "--disable-dev-shm-usage",
            ],
        )
        ctx = await browser.new_context(
            viewport={"width": 1440, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-US",
            timezone_id="America/New_York",
        )
        await ctx.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
        )
        page = await ctx.new_page()

        try:
            await page.goto(
                "https://wellfound.com/role/r/backend-developer?remote=true",
                wait_until="domcontentloaded",
                timeout=45000,
            )
            await page.wait_for_timeout(4000)

            # Scroll to load lazy content
            for _ in range(6):
                await page.keyboard.press("End")
                await page.wait_for_timeout(random.randint(1200, 2000))

            content = await page.content()

            # Attempt __NEXT_DATA__ extraction
            nd_match = re.search(r'id="__NEXT_DATA__"[^>]*>(.*?)</script>', content, re.DOTALL)
            if nd_match:
                try:
                    nd = json.loads(nd_match.group(1))

                    def _walk(obj: object, depth: int = 0):
                        if depth > 12 or len(raw_jobs) >= max_jobs:
                            return
                        if isinstance(obj, dict):
                            job_url = obj.get("jobUrl") or obj.get("applyUrl") or ""
                            if job_url and job_url not in seen:
                                if not job_url.startswith("http"):
                                    job_url = "https://wellfound.com" + job_url
                                seen.add(job_url)
                                desc = obj.get("description", "")
                                raw_jobs.append({
                                    "title":       obj.get("jobTitle") or obj.get("title", ""),
                                    "company":     obj.get("companyName") or obj.get("company", ""),
                                    "url":         job_url,
                                    "location":    obj.get("locationName") or obj.get("location", ""),
                                    "description": html_to_text(desc) if "<" in desc else desc,
                                })
                            for v in list(obj.values())[:WELLFOUND_WALK_MAX_BREADTH]:
                                _walk(v, depth + 1)
                        elif isinstance(obj, list):
                            for item in obj[:WELLFOUND_WALK_MAX_BREADTH]:
                                _walk(item, depth + 1)

                    _walk(nd)
                except Exception:
                    pass

            # Fallback: parse job href patterns
            if not raw_jobs:
                for link in re.findall(r'href="(/jobs/[^"?#]+)"', content):
                    if len(raw_jobs) >= max_jobs:
                        break
                    full_url = f"https://wellfound.com{link}"
                    if full_url in seen:
                        continue
                    seen.add(full_url)
                    slug = link.rstrip("/").split("/")[-1]
                    raw_jobs.append({
                        "title":       slug.replace("-", " ").title(),
                        "company":     "",
                        "url":         full_url,
                        "description": "",
                    })

        except Exception as e:
            print(f"  [WARN] Wellfound Playwright error: {type(e).__name__}: {e}")
        finally:
            await browser.close()

    return raw_jobs[:max_jobs]


def scrape_wellfound(max_jobs: int = MAX_JOBS) -> list[dict]:
    print("\n[GEMA] Scraping Wellfound (Playwright stealth)...")
    try:
        raw_jobs = asyncio.run(_playwright_wellfound(max_jobs))
    except Exception as e:
        print(f"  [ERROR] Wellfound: {e}")
        raw_jobs = []
    print(f"  Found {len(raw_jobs)} Wellfound jobs")
    return raw_jobs


# =============================================================================
# SCRAPER 8: WelcomeToTheJungle (Playwright)
# =============================================================================

async def _playwright_wttj(max_jobs: int) -> list[dict]:
    from playwright.async_api import async_playwright  # type: ignore

    raw_jobs: list[dict] = []
    seen: set[str] = set()

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        ctx = await browser.new_context(
            viewport={"width": 1440, "height": 900},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-US",
        )
        page = await ctx.new_page()

        try:
            await page.goto(
                "https://www.welcometothejungle.com/en/jobs"
                "?refinementList%5Bremote%5D%5B0%5D=fulltime"
                "&query=backend+engineer",
                wait_until="domcontentloaded",
                timeout=45000,
            )
            await page.wait_for_timeout(4000)

            for _ in range(6):
                await page.keyboard.press("End")
                await page.wait_for_timeout(random.randint(1500, 2200))

            content = await page.content()

            # WTTJ jobs are at /en/companies/{slug}/jobs/{job-slug}
            job_links = list(set(re.findall(
                r'href="(/en/companies/[^/]+/jobs/[^"?#]+)"', content
            )))

            # Try Apollo state for richer data
            apollo_match = re.search(
                r'window\.__APOLLO_STATE__\s*=\s*(\{.+?\});\s*</script>',
                content, re.DOTALL
            )
            if apollo_match:
                try:
                    apollo = json.loads(apollo_match.group(1))
                    for val in apollo.values():
                        if not isinstance(val, dict) or len(raw_jobs) >= max_jobs:
                            continue
                        if val.get("__typename") != "Job":
                            continue
                        job_url = val.get("websiteUrl") or ""
                        slug = val.get("slug", "")
                        if not job_url and slug:
                            job_url = f"https://www.welcometothejungle.com/en/jobs/{slug}"
                        if not job_url or job_url in seen:
                            continue
                        seen.add(job_url)
                        org = val.get("organization") or {}
                        desc = val.get("description", "")
                        raw_jobs.append({
                            "title":       val.get("name", ""),
                            "company":     org.get("name", "") if isinstance(org, dict) else "",
                            "url":         job_url,
                            "location":    "",
                            "description": html_to_text(desc) if "<" in desc else desc,
                        })
                except Exception:
                    pass

            # Fallback: use discovered href links
            for link in job_links:
                if len(raw_jobs) >= max_jobs:
                    break
                full_url = f"https://www.welcometothejungle.com{link}"
                if full_url in seen:
                    continue
                seen.add(full_url)
                parts = link.strip("/").split("/")
                company_slug = parts[2] if len(parts) > 2 else ""
                job_slug = parts[-1] if len(parts) > 4 else ""
                raw_jobs.append({
                    "title":       job_slug.replace("-", " ").title(),
                    "company":     company_slug.replace("-", " ").title(),
                    "url":         full_url,
                    "description": "",
                })

        except Exception as e:
            print(f"  [WARN] WTTJ Playwright error: {type(e).__name__}: {e}")
        finally:
            await browser.close()

    return raw_jobs[:max_jobs]


def scrape_welcometothejungle(max_jobs: int = MAX_JOBS) -> list[dict]:
    print("\n[GEMA] Scraping WelcomeToTheJungle (Playwright)...")
    try:
        raw_jobs = asyncio.run(_playwright_wttj(max_jobs))
    except Exception as e:
        print(f"  [ERROR] WelcomeToTheJungle: {e}")
        raw_jobs = []
    print(f"  Found {len(raw_jobs)} WelcomeToTheJungle jobs")
    return raw_jobs


# =============================================================================
# Excel writer — multi-tab, sorted, formatted
# =============================================================================

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FONT  = Font(color="276221")
RED_FONT    = Font(color="9C0006")


def _format_sheet(ws, df: pd.DataFrame) -> None:
    if ws.max_row < 1:
        return

    col_map = {name: get_column_letter(i + 1) for i, name in enumerate(df.columns)}

    # Headers
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # MATCH_SCORE formatting + conditional colour
    mc = col_map.get("MATCH_SCORE")
    if mc and ws.max_row > 1:
        score_range = f"{mc}2:{mc}{ws.max_row}"
        score_col_idx = df.columns.get_loc("MATCH_SCORE") + 1
        for row in ws.iter_rows(min_row=2, min_col=score_col_idx, max_col=score_col_idx):
            for cell in row:
                cell.number_format = "0%"
                cell.alignment = Alignment(horizontal="center")
        ws.conditional_formatting.add(
            score_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["0.7"],
                       fill=GREEN_FILL, font=GREEN_FONT),
        )
        ws.conditional_formatting.add(
            score_range,
            CellIsRule(operator="lessThan", formula=["0.3"],
                       fill=RED_FILL, font=RED_FONT),
        )

    # APPLY_LINK — clickable blue hyperlinks
    lc = col_map.get("APPLY_LINK")
    if lc and ws.max_row > 1:
        link_col_idx = df.columns.get_loc("APPLY_LINK") + 1
        for row in ws.iter_rows(min_row=2, min_col=link_col_idx, max_col=link_col_idx):
            for cell in row:
                url = str(cell.value or "")
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.value     = url
                    cell.font      = Font(color="0563C1", underline="single")

    # General cell alignment
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter != lc:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Auto-width
    MIN_W, MAX_W = 12, 60
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
            default=MIN_W,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, MIN_W), MAX_W)

    # Auto-filter + freeze top row
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"

    # Row heights
    ws.row_dimensions[1].height = 22
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 55


def write_multitab_excel(sheet_data: dict, xlsx_path: Path) -> None:
    """
    Merge new sheet_data into the existing Excel file (all sheets preserved).
    Each sheet is sorted by MATCH_SCORE descending before saving.
    """
    # Load all existing sheets
    existing: dict[str, pd.DataFrame] = {}
    if xlsx_path.exists():
        try:
            with pd.ExcelFile(xlsx_path, engine="openpyxl") as xf:
                for sname in xf.sheet_names:
                    existing[sname] = pd.read_excel(xf, sheet_name=sname)
        except Exception as e:
            print(f"  [WARN] Could not load existing Excel: {e}")

    # Merge new data into correct sheets
    merged: dict[str, pd.DataFrame] = {}
    for sname, new_df in sheet_data.items():
        if sname in existing and not existing[sname].empty:
            combined = pd.concat([existing[sname], new_df], ignore_index=True)
        else:
            combined = new_df
        if "MATCH_SCORE" in combined.columns:
            combined = combined.sort_values("MATCH_SCORE", ascending=False).reset_index(drop=True)
        merged[sname] = combined

    # Preserve sheets not updated this run
    for sname, df in existing.items():
        if sname not in merged:
            merged[sname] = df

    # Write all sheets to disk
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        for sname, df in merged.items():
            safe_name = sname[:31]  # Excel tab name max 31 chars
            df.to_excel(writer, sheet_name=safe_name, index=False)

    # Apply per-sheet formatting
    wb = load_workbook(xlsx_path)
    for sname, df in merged.items():
        safe_name = sname[:31]
        if safe_name in wb.sheetnames:
            _format_sheet(wb[safe_name], df)
    wb.save(xlsx_path)


# =============================================================================
# Source registry + main
# =============================================================================

SOURCES = [
    {"name": "HackerNews",        "fn": scrape_hackernews,        "src_url": "https://news.ycombinator.com/jobs"},
    {"name": "WorkingNomads",     "fn": scrape_workingnomads,     "src_url": "https://www.workingnomads.com/jobs"},
    {"name": "Himalayas",         "fn": scrape_himalayas,         "src_url": "https://himalayas.app/jobs/"},
    {"name": "Remotivated",       "fn": scrape_remotivated,       "src_url": "https://remotivated.com/jobs"},
    {"name": "PostHog",           "fn": scrape_posthog,           "src_url": "https://posthog.com/cool-tech-jobs"},
    {"name": "Greenhouse",        "fn": scrape_greenhouse,        "src_url": "https://www.greenhouse.com/careers/opportunities"},
    {"name": "Wellfound",         "fn": scrape_wellfound,         "src_url": "https://wellfound.com/jobs"},
    {"name": "WelcomeToTheJungle","fn": scrape_welcometothejungle,"src_url": "https://app.welcometothejungle.com/jobs/"},
]


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="GEMA Industrial Multi-Source Scraper")
    parser.add_argument(
        "--cv",
        type=str,
        default=None,
        metavar="PATH",
        help="Path to CV file (PDF/DOCX/TXT/MD) — enables profile-aware NLP extraction.",
    )
    args = parser.parse_args()

    extraction_prompt: Optional[str] = None
    if args.cv:
        cv_path = Path(args.cv)
        if cv_path.exists():
            file_bytes  = cv_path.read_bytes()
            raw_cv_text = extract_cv_text(cv_path.name, file_bytes)
            profile     = generate_ephemeral_profile(raw_cv_text)
            extraction_prompt = _build_extraction_prompt(profile)
            print(
                f"[GEMA] CV loaded: {cv_path.name} — "
                f"role={profile.get('role', '?')}, "
                f"skills={len(profile.get('core_skills', []))}, "
                f"signals={len(profile.get('audit_signals', []))}"
            )
        else:
            print(f"[WARN] CV file not found: {args.cv} — using generic extraction profile.")

    reset_rate_limit_flags()

    total_new  = 0
    sheet_data: dict[str, pd.DataFrame] = {}
    seen_urls:  set[str] = set()

    for source in SOURCES:
        name    = source["name"]
        fn      = source["fn"]
        src_url = source["src_url"]

        try:
            raw_jobs = fn(MAX_JOBS)
        except Exception as e:
            print(f"  [ERROR] {name} scraper raised: {e}")
            raw_jobs = []

        if not raw_jobs:
            print(f"  [SKIP] {name}: 0 jobs collected")
            continue

        print(f"\n[GEMA] Extracting {len(raw_jobs)} jobs from {name}...")
        rows = extract_and_build_rows(
            raw_jobs,
            src_url,
            seen_urls=seen_urls,
            extraction_prompt=extraction_prompt,
        )

        if rows:
            sheet_data[name] = pd.DataFrame(rows, columns=COLUMNS)
            total_new += len(rows)
            print(f"  [OK] {name}: {len(rows)} jobs extracted")
        else:
            print(f"  [SKIP] {name}: NLP returned 0 valid jobs")

    if sheet_data:
        print(f"\n[GEMA] Writing {total_new} new jobs to Excel (multi-tab)...")
        write_multitab_excel(sheet_data, XLSX_PATH)
        print(f"\n📈 GEMA Industrial Tracker Updated. Total new jobs analyzed: {total_new}")
        print(f"   {XLSX_PATH}")
    else:
        print("\n[GEMA] No valid jobs extracted. Excel unchanged.")
