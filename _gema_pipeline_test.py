"""
_gema_pipeline_test.py — GEMA Personalized Extraction Pipeline
Fetches G2i jobs, runs NLP extraction, writes GEMA_JOB_TRACKER.xlsx.
"""

import json
import sys
import urllib.request
from datetime import date
from html.parser import HTMLParser
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, ".")
from dotenv import load_dotenv
load_dotenv()

from nlp_engine import extract_jobs_from_text

ASHBY_API  = "https://api.ashbyhq.com/posting-api/job-board/g2i?includeCompensation=true"
SOURCE_URL = "https://jobs.ashbyhq.com/g2i?locationId=7b7cca1a-2f42-4ee6-8028-63b53269f337&workplaceType=Remote"
XLSX_PATH  = Path(__file__).parent / "GEMA_JOB_TRACKER.xlsx"
BATCH_SIZE = 5

COLUMNS = [
    "DATE_FETCHED", "MATCH_SCORE", "JOB_TITLE", "COMPANY",
    "SALARY", "LOCATION_STRICTNESS", "TECH_STACK", "APPLY_LINK", "NOTES",
]


# =============================================================================
# HTML stripper
# =============================================================================

class _Strip(HTMLParser):
    def __init__(self):
        super().__init__()
        self._p = []
    def handle_data(self, d):
        t = d.strip()
        if t:
            self._p.append(t)
    def get_text(self):
        return "\n".join(self._p)

def html_to_text(h: str) -> str:
    p = _Strip()
    p.feed(h.encode("utf-8", errors="replace").decode("utf-8", errors="replace"))
    return p.get_text()


# =============================================================================
# Field formatters
# =============================================================================

def fmt_salary(job) -> str:
    if job.salary_min is None and job.salary_max is None:
        return "Not disclosed"
    curr = job.currency or "USD"
    lo = f"{job.salary_min:,}" if job.salary_min else "?"
    hi = f"{job.salary_max:,}" if job.salary_max else "?"
    if job.salary_min and job.salary_max and job.salary_min != job.salary_max:
        return f"{curr} {lo} – {hi}"
    return f"{curr} {lo}"

def fmt_tech(job) -> str:
    return ", ".join(job.required_tech) if job.required_tech else "Not specified"

def build_block(i: int, job: dict) -> tuple[str, str]:
    comp    = job.get("compensation") or {}
    smin    = comp.get("minValue")
    smax    = comp.get("maxValue")
    curr    = comp.get("currency", "USD")
    sal     = (f"Salary: {curr} {smin:,} - {smax:,}\n" if smin and smax
               else (f"Salary: {curr} {smin:,}+\n" if smin else ""))
    desc    = html_to_text(job.get("descriptionHtml") or job.get("descriptionSocial") or "")
    team    = job.get("team") or {}
    dept    = team.get("name", "") if isinstance(team, dict) else ""
    company = dept if dept else "G2i"
    url     = job.get("jobUrl") or SOURCE_URL
    block   = (
        f"Job {i}:\nTitle: {job.get('title','')}\nCompany: {company}\n"
        f"Location: {job.get('location','')}\n{sal}"
        f"URL: {url}\n\n{desc}"
    )
    return url, block


# =============================================================================
# Excel writer with full formatting
# =============================================================================

HEADER_FILL  = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
GREEN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FONT   = Font(color="276221")
RED_FONT     = Font(color="9C0006")

def write_excel(df: pd.DataFrame, xlsx_path: Path) -> None:
    """Save DataFrame to Excel with full formatting."""
    df.to_excel(xlsx_path, index=False, engine="openpyxl")
    wb = load_workbook(xlsx_path)
    ws = wb.active

    col_map = {name: get_column_letter(i + 1) for i, name in enumerate(df.columns)}

    # ── Header row ────────────────────────────────────────────────────────────
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # ── MATCH_SCORE: percentage format + conditional colour ───────────────────
    mc = col_map.get("MATCH_SCORE")
    if mc:
        score_range = f"{mc}2:{mc}{ws.max_row}"
        for row in ws.iter_rows(min_row=2, min_col=df.columns.get_loc("MATCH_SCORE") + 1,
                                max_col=df.columns.get_loc("MATCH_SCORE") + 1):
            for cell in row:
                cell.number_format = "0%"
                cell.alignment = Alignment(horizontal="center")

        ws.conditional_formatting.add(
            score_range,
            CellIsRule(operator="greaterThanOrEqual", formula=["0.7"],
                       fill=GREEN_FILL, font=GREEN_FONT),
        )
        ws.conditional_formatting.add(
            score_range,
            CellIsRule(operator="lessThan", formula=["0.3"],
                       fill=RED_FILL, font=RED_FONT),
        )

    # ── APPLY_LINK: blue clickable hyperlinks ─────────────────────────────────
    lc = col_map.get("APPLY_LINK")
    if lc:
        link_col_idx = df.columns.get_loc("APPLY_LINK") + 1
        for row in ws.iter_rows(min_row=2, min_col=link_col_idx, max_col=link_col_idx):
            for cell in row:
                url = str(cell.value or "")
                if url.startswith("http"):
                    cell.hyperlink = url
                    cell.value     = url
                    cell.font      = Font(color="0563C1", underline="single")

    # ── General cell alignment ─────────────────────────────────────────────────
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter != lc:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Auto-width ─────────────────────────────────────────────────────────────
    MIN_W, MAX_W = 12, 60
    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
            default=MIN_W,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, MIN_W), MAX_W)

    # ── Auto-filter + freeze top row ──────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes    = "A2"

    # ── Row height for data rows ───────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 55

    wb.save(xlsx_path)


# =============================================================================
# Fetch jobs
# =============================================================================

print("[GEMA] Fetching from Ashby API...")
req = urllib.request.Request(
    ASHBY_API,
    headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"},
)
with urllib.request.urlopen(req, timeout=15) as r:
    all_jobs = json.loads(r.read().decode("utf-8")).get("jobs", [])

print(f"[GEMA] {len(all_jobs)} jobs found. Processing in batches of {BATCH_SIZE}...\n")

today     = date.today().isoformat()
new_rows  = []

# =============================================================================
# Extract in batches
# =============================================================================

for batch_start in range(0, len(all_jobs), BATCH_SIZE):
    batch   = all_jobs[batch_start: batch_start + BATCH_SIZE]
    blocks  = []
    url_map = {}

    for local_i, job in enumerate(batch, 1):
        global_i         = batch_start + local_i
        url, block       = build_block(global_i, job)
        url_map[global_i] = url
        blocks.append(block)

    raw = "\n\n" + ("\n\n" + "-" * 40 + "\n\n").join(blocks)
    result = extract_jobs_from_text(raw, source_url=SOURCE_URL, log_callback=None)

    for job in result.jobs:
        url = job.source_url or SOURCE_URL
        new_rows.append({
            "DATE_FETCHED":       today,
            "MATCH_SCORE":        job.cv_match_score,
            "JOB_TITLE":          job.title,
            "COMPANY":            job.company,
            "SALARY":             fmt_salary(job),
            "LOCATION_STRICTNESS": job.location_strictness or "Unknown",
            "TECH_STACK":         fmt_tech(job),
            "APPLY_LINK":         url,
            "NOTES":              job.location_notes or "",
        })
        print(f"  📈 Processed: {job.title} at {job.company} - Saved to Excel.")

# =============================================================================
# Append logic — load existing, merge, save
# =============================================================================

if XLSX_PATH.exists():
    existing = pd.read_excel(XLSX_PATH, engine="openpyxl")
    df = pd.concat([existing, pd.DataFrame(new_rows, columns=COLUMNS)], ignore_index=True)
else:
    df = pd.DataFrame(new_rows, columns=COLUMNS)

write_excel(df, XLSX_PATH)

print(f"\n📈 GEMA Tracker Updated: GEMA_JOB_TRACKER.xlsx [Added {len(new_rows)} new jobs]")
print(f"   {XLSX_PATH}")
